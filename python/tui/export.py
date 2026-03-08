"""
Export snapshot and box spread data to CSV and Excel for use in Google Sheets / Excel.

Use from TUI (Export action, e.g. F6) or programmatically. Files are written with
timestamped names so each export is preserved. CSV opens in both Excel and Google Sheets
(File > Import or drag-and-drop). For live sync to Google Sheets, point Sheets to
a recurring export path or use Google Sheets API (see docs).
"""

from __future__ import annotations

import csv
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, List, Dict, Any

from .models import (
    SnapshotPayload,
    BoxSpreadPayload,
    PositionSnapshot,
    FutureEvent,
    BoxSpreadScenario,
)


def _export_dir() -> Path:
    """Default export directory: build/export, or EXPORT_DIR env."""
    env = os.environ.get("TUI_EXPORT_DIR")
    if env:
        return Path(env).expanduser().resolve()
    root = Path(__file__).resolve().parents[2]  # repo root from python/tui/
    return (root / "build" / "export").resolve()


def _timestamp_prefix() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%SZ")


def _position_row(p: PositionSnapshot) -> Dict[str, Any]:
    """Flatten a position into a single row for CSV/Excel."""
    d: Dict[str, Any] = {
        "name": p.name,
        "quantity": p.quantity,
        "roi": p.roi,
        "instrument_type": p.instrument_type or "",
        "currency": p.currency or "USD",
        "market_value": p.market_value,
        "rate": p.rate,
        "maturity_date": p.maturity_date or "",
        "cash_flow": p.cash_flow,
        "side": p.side or "",
        "price": p.price,
        "bid": p.bid,
        "ask": p.ask,
        "last": p.last,
        "spread": p.spread,
        "expected_cash_at_expiry": p.expected_cash_at_expiry,
        "dividend": p.dividend,
    }
    return d


def _future_event_row(e: FutureEvent) -> Dict[str, Any]:
    return {
        "event_type": e.event_type,
        "date": e.date,
        "amount": e.amount,
        "currency": e.currency,
        "position_name": e.position_name,
        "description": e.description,
    }


def _scenario_row(s: BoxSpreadScenario) -> Dict[str, Any]:
    return {
        "width": s.width,
        "put_bid": s.put_bid,
        "call_ask": s.call_ask,
        "synthetic_bid": s.synthetic_bid,
        "synthetic_ask": s.synthetic_ask,
        "mid_price": s.mid_price,
        "annualized_return": s.annualized_return,
        "fill_probability": s.fill_probability,
        "option_style": s.option_style,
        "expiration_date": s.expiration_date or "",
        "days_to_expiry": s.days_to_expiry,
        "buy_profit": s.buy_profit,
        "sell_profit": s.sell_profit,
    }


def export_snapshot_to_csv(
    snapshot: SnapshotPayload,
    dir_path: Optional[Path] = None,
    prefix: Optional[str] = None,
) -> List[Path]:
    """
    Write snapshot positions and future_events to CSV files.
    Returns paths to the written files.
    """
    export_path = dir_path or _export_dir()
    export_path.mkdir(parents=True, exist_ok=True)
    pre = prefix or _timestamp_prefix()
    written: List[Path] = []

    if snapshot.positions:
        positions_path = export_path / f"{pre}_positions.csv"
        keys = list(_position_row(snapshot.positions[0]).keys())
        with open(positions_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
            w.writeheader()
            for p in snapshot.positions:
                w.writerow(_position_row(p))
        written.append(positions_path)

    if snapshot.future_events:
        events_path = export_path / f"{pre}_future_events.csv"
        keys = list(_future_event_row(snapshot.future_events[0]).keys())
        with open(events_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
            w.writeheader()
            for e in snapshot.future_events:
                w.writerow(_future_event_row(e))
        written.append(events_path)

    return written


def export_box_spread_to_csv(
    payload: BoxSpreadPayload,
    dir_path: Optional[Path] = None,
    prefix: Optional[str] = None,
) -> Optional[Path]:
    """Write box spread scenarios to a CSV file. Returns path or None if no scenarios."""
    if not payload.scenarios:
        return None
    export_path = dir_path or _export_dir()
    export_path.mkdir(parents=True, exist_ok=True)
    pre = prefix or _timestamp_prefix()
    path = export_path / f"{pre}_box_spread_scenarios.csv"
    keys = list(_scenario_row(payload.scenarios[0]).keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        for s in payload.scenarios:
            w.writerow(_scenario_row(s))
    return path


def export_snapshot_to_xlsx(
    snapshot: SnapshotPayload,
    dir_path: Optional[Path] = None,
    prefix: Optional[str] = None,
) -> Optional[Path]:
    """Write snapshot positions and future_events to a single Excel file. Returns path."""
    try:
        import openpyxl
    except ImportError:
        return None

    export_path = dir_path or _export_dir()
    export_path.mkdir(parents=True, exist_ok=True)
    pre = prefix or _timestamp_prefix()
    path = export_path / f"{pre}_snapshot.xlsx"
    wb = openpyxl.Workbook()

    if snapshot.positions:
        ws = wb.active
        ws.title = "Positions"
        rows = [_position_row(p) for p in snapshot.positions]
        keys = list(rows[0].keys()) if rows else []
        for c, k in enumerate(keys, 1):
            ws.cell(row=1, column=c, value=k)
        for r, row in enumerate(rows, 2):
            for c, k in enumerate(keys, 1):
                ws.cell(row=r, column=c, value=row.get(k))

    if snapshot.future_events:
        ws2 = wb.create_sheet("Future events")
        rows = [_future_event_row(e) for e in snapshot.future_events]
        keys = list(rows[0].keys()) if rows else []
        for c, k in enumerate(keys, 1):
            ws2.cell(row=1, column=c, value=k)
        for r, row in enumerate(rows, 2):
            for c, k in enumerate(keys, 1):
                ws2.cell(row=r, column=c, value=row.get(k))

    wb.save(path)
    return path


def export_box_spread_to_xlsx(
    payload: BoxSpreadPayload,
    dir_path: Optional[Path] = None,
    prefix: Optional[str] = None,
) -> Optional[Path]:
    """Write box spread scenarios to an Excel file. Returns path or None."""
    if not payload.scenarios:
        return None
    try:
        import openpyxl
    except ImportError:
        return None

    export_path = dir_path or _export_dir()
    export_path.mkdir(parents=True, exist_ok=True)
    pre = prefix or _timestamp_prefix()
    path = export_path / f"{pre}_box_spread.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scenarios"
    rows = [_scenario_row(s) for s in payload.scenarios]
    keys = list(rows[0].keys())
    for c, k in enumerate(keys, 1):
        ws.cell(row=1, column=c, value=k)
    for r, row in enumerate(rows, 2):
        for c, k in enumerate(keys, 1):
            ws.cell(row=r, column=c, value=row.get(k))
    wb.save(path)
    return path


def export_all(
    snapshot: Optional[SnapshotPayload],
    box_spread: Optional[BoxSpreadPayload],
    dir_path: Optional[Path] = None,
    formats: Optional[List[str]] = None,
) -> List[Path]:
    """
    Export snapshot and box spread to CSV (and optionally xlsx).
    formats: ["csv"] or ["csv", "xlsx"] (default: both).
    Returns list of written file paths.
    """
    formats = formats or ["csv", "xlsx"]
    export_path = dir_path or _export_dir()
    prefix = _timestamp_prefix()
    written: List[Path] = []

    if snapshot:
        if "csv" in formats:
            written.extend(
                export_snapshot_to_csv(snapshot, dir_path=export_path, prefix=prefix)
            )
        if "xlsx" in formats:
            p = export_snapshot_to_xlsx(
                snapshot, dir_path=export_path, prefix=prefix
            )
            if p:
                written.append(p)

    if box_spread:
        if "csv" in formats:
            p = export_box_spread_to_csv(
                box_spread, dir_path=export_path, prefix=prefix
            )
            if p:
                written.append(p)
        if "xlsx" in formats:
            p = export_box_spread_to_xlsx(
                box_spread, dir_path=export_path, prefix=prefix
            )
            if p:
                written.append(p)

    return written
